"""
excel_parser.py - BM기획서 및 매출지표 Excel 파서
openpyxl 기반 브라우저 독립적 파서
"""
import re
import openpyxl
from io import BytesIO

# 시뮬레이터 아이템 매핑 (substring 매칭)
SIMULATOR_ITEMS = [
    '찬란한 클래스 11회', '신비로운 클래스 11회', '눈부신 클래스 11회', '영롱한 클래스 11회',
    '찬란한 펫 11회', '신비로운 펫 11회', '눈부신 펫 11회', '영롱한 펫 11회',
    '찬란한 투혼 11회', '신비로운 투혼 11회', '눈부신 투혼 11회', '영롱한 투혼 11회',
    '찬란한 카드 11회', '신비로운 카드 11회', '눈부신 카드 11회',
    '영웅 클래스 도전', '고대 클래스 도전', '전설 클래스 도전',
    '영웅 펫 도전', '고대 펫 도전', '전설 펫 도전',
    '영웅 투혼 도전', '고대 투혼 도전', '전설 투혼 도전',
    '영웅 카드 도전', '고대 카드 도전', '전설 카드 도전',
    '영웅 클래스 확정', '고대 클래스 확정', '전설 클래스 확정',
    '영웅 펫 확정', '고대 펫 확정', '전설 펫 확정',
    '영웅 투혼 확정', '고대 투혼 확정', '전설 투혼 확정',
    '영웅 카드 확정', '고대 카드 확정', '전설 카드 확정',
    '미숙한 보물사냥꾼', '숙달된 보물사냥꾼', '능숙한 보물사냥꾼',
    '노련한 보물사냥꾼', '대단한 보물사냥꾼', '완벽한 보물사냥꾼',
    '노련한 각성자', '대단한 각성자', '완벽한 각성자',
    '노련한 동반자', '대단한 동반자', '완벽한 동반자',
]


def match_item(raw_name):
    """아이템 이름 매칭"""
    if not raw_name:
        return None
    s = str(raw_name)
    for keyword in SIMULATOR_ITEMS:
        if keyword in s:
            return keyword
    return None


def is_garbage_pkg(name):
    if not name:
        return True
    n = str(name).strip()
    return n == '패키지 명' or n.startswith('PackageBox_') or n == '천사의 미소'


def extract_date_label(filename):
    """파일명에서 날짜 추출: '0311BM.xlsx' → ('0311', '3월 11일 기획서')"""
    m = re.match(r'^(\d{2})(\d{2})', filename)
    if m:
        mm, dd = int(m.group(1)), int(m.group(2))
        return m.group(1) + m.group(2), f'{mm}월 {dd}일 기획서'
    base = re.sub(r'\.xlsx?$', '', filename, flags=re.IGNORECASE)
    return f'etc_{base}', base


def parse_bm_excel(file_bytes, filename):
    """BM기획서 Excel 파싱 → 패키지 리스트"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    if '입력' not in wb.sheetnames:
        return []
    ws = wb['입력']

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    # 헤더 자동 감지
    col_pkg, col_item, col_qty, col_price, col_limit = -1, -1, -1, -1, -1
    header_row = -1

    for r in range(min(5, len(rows))):
        row = rows[r]
        for c, v in enumerate(row):
            vs = str(v or '')
            if '패키지 명' in vs or vs == '패키지명':
                col_pkg = c
            if vs == '가격':
                col_price = c
            if vs in ('구매 제한', '구매제한'):
                col_limit = c
        if col_pkg >= 0:
            header_row = r
            break

    if header_row >= 0 and header_row + 1 < len(rows):
        sub_row = rows[header_row + 1]
        for c, v in enumerate(sub_row):
            vs = str(v or '')
            if vs == '아이템':
                col_item = c
            if vs == '수량':
                col_qty = c

    # 폴백
    if col_pkg < 0: col_pkg = 4
    if col_item < 0: col_item = col_pkg + 2
    if col_qty < 0: col_qty = col_item + 1
    if col_price < 0: col_price = col_pkg + 7
    if col_limit < 0: col_limit = col_pkg + 5

    data_start = header_row + 2 if header_row >= 0 else 3
    date_key, _ = extract_date_label(filename)

    packages = []
    current_pkg = None
    pkg_idx = 0

    for i in range(data_start, len(rows)):
        row = rows[i]
        if not row:
            continue

        def safe_get(idx):
            return row[idx] if idx < len(row) else None

        pkg_name = safe_get(col_pkg)
        price = safe_get(col_price)
        item_raw = safe_get(col_item)
        qty = safe_get(col_qty)
        limit = safe_get(col_limit)

        if pkg_name is not None:
            if not is_garbage_pkg(pkg_name):
                pkg_idx += 1
                current_pkg = {
                    'id': f'pkg_{date_key}_{pkg_idx:03d}',
                    'name': f'[{date_key}] {str(pkg_name).strip()}',
                    'price': price if price else None,
                    'purchaseLimit': str(limit).strip() if limit else None,
                    'items': [],
                    'rawItems': [],
                }
                packages.append(current_pkg)
            else:
                current_pkg = None

        if current_pkg and item_raw is not None:
            raw_name = str(item_raw).strip()
            try:
                q = int(qty) if qty else 0
            except (ValueError, TypeError):
                q = 0

            if raw_name and q > 0:
                current_pkg['rawItems'].append({'name': raw_name, 'quantity': q})

            sim_name = match_item(item_raw)
            if sim_name and qty is not None:
                existing = next((it for it in current_pkg['items'] if it['name'] == sim_name), None)
                if existing:
                    existing['quantity'] += q
                else:
                    current_pkg['items'].append({'name': sim_name, 'quantity': q})

    return [p for p in packages if len(p['items']) > 0]


def parse_sales_excel(file_bytes):
    """매출지표 Excel 파싱"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    # 기간 추출
    period_label = ''
    for r in range(min(8, len(rows))):
        row = rows[r]
        if not row:
            continue
        for cell in row:
            if not cell:
                continue
            m = re.search(r'(\d{4}-\d{2}-\d{2})\s*~\s*(\d{4}-\d{2}-\d{2})', str(cell))
            if m:
                period_label = f'{m.group(1)} ~ {m.group(2)}'
                break
        if period_label:
            break

    # 헤더 감지
    header_row_idx = -1
    col_item, col_revenue, col_buyers, col_purchases = -1, -1, -1, -1

    for r in range(min(15, len(rows))):
        row = rows[r]
        if not row:
            continue
        for c, v in enumerate(row):
            vs = str(v or '')
            if '매출아이템명' in vs:
                header_row_idx = r
                col_item = c
                for cc, vv in enumerate(row):
                    h = str(vv or '')
                    if '개인매출' in h:
                        col_revenue = cc
                    elif '구매유저' in h:
                        col_buyers = cc
                    elif '구매횟수' in h:
                        col_purchases = cc
                break
        if header_row_idx >= 0:
            break

    if header_row_idx < 0:
        return None

    items = {}
    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        if not row:
            continue
        def safe_get(idx):
            return row[idx] if idx < len(row) else None

        item_name = safe_get(col_item)
        if not item_name:
            continue
        name = str(item_name).strip()

        try:
            revenue = float(safe_get(col_revenue) or 0)
        except (ValueError, TypeError):
            revenue = 0
        try:
            buyers = float(safe_get(col_buyers) or 0)
        except (ValueError, TypeError):
            buyers = 0
        try:
            purchases = float(safe_get(col_purchases) or 0)
        except (ValueError, TypeError):
            purchases = 0

        if name in items:
            items[name]['revenue'] += revenue
            items[name]['buyers'] = max(items[name]['buyers'], buyers)
            items[name]['purchases'] += purchases
        else:
            items[name] = {'revenue': revenue, 'buyers': buyers, 'purchases': purchases}

    return {'periodLabel': period_label, 'items': items}
